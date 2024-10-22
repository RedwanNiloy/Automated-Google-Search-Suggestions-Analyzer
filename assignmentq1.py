# -*- coding: utf-8 -*-
"""AssignmentQ1.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/15hyNw5jZQ5VlQt4X05HItXqrzO0DKGCP
"""

pip install selenium openpyxl

#Use google collab to run this srcipt

pip install selenium openpyxl

import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl
from datetime import datetime

# Installed Chromium and ChromeDriver in Colab
!apt-get update
!apt install -y chromium-chromedriver

# Setting up Chrome options for Colab
chrome_options = Options()
chrome_options.add_argument("--headless")  # Run Chrome in headless mode
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# Initializing the WebDriver
driver = webdriver.Chrome(options=chrome_options)

# Loading the Excel workbook
workbook_path = "/content/sample_data/4BeatsQ1.xlsx"  # Replace with your file path in google collab
workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook.active

# Iterating over each keyword in the Excel sheet
for row in range(3, 13):        #Starting from row 3 to row 12
    keyword = sheet.cell(row=row, column=3).value  #started from column 3 as it is the column of keywords
    print(keyword)

    # Going to Google and searching the keyword
    driver.get("https://www.google.com")
    search_box = driver.find_element("name", "q")
    search_box.clear()
    search_box.send_keys(keyword)
    time.sleep(2)  # Waiting for suggestions to load

    # Getting suggestions using XPath
    suggestions = driver.find_elements("xpath", "//ul[@role='listbox']//li[@role='presentation']//span")
    suggestion_texts = [s.text for s in suggestions if s.text]

    if suggestion_texts:
        # Finding the longest and shortest suggestion
        longest_suggestion = max(suggestion_texts, key=len)
        shortest_suggestion = min(suggestion_texts, key=len)
    else:
        longest_suggestion = "No suggestions"
        shortest_suggestion = "No suggestions"

    # Getting current date and time for each search
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Writing the results back to the Excel sheet
    sheet.cell(row=row, column=4).value = longest_suggestion
    sheet.cell(row=row, column=5).value = shortest_suggestion
    sheet.cell(row=row, column=6).value = current_time


    workbook.save(workbook_path)

driver.quit()

workbook.save(workbook_path)